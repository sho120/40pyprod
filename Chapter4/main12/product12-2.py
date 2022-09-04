from openpyxl import load_workbook

load_wb = load_workbook(r'Chapter4\main12\certificate.xlsx')
load_ws = load_wb.active

name_list = []
birth_list = []
ho_list = []

for i in range(1,load_ws.max_row+1):
    name_list.append(load_ws.cell(i,1).value)
    birth_list.append(load_ws.cell(i,2).value)
    ho_list.append(load_ws.cell(i,3).value)

print(name_list)
print(birth_list)
print(ho_list)